import os
import uuid
import json
import shutil
import datetime
import numpy as np
from typing import List
from fastapi import APIRouter, UploadFile, File, HTTPException
from redis import Redis
from redisvl.index import SearchIndex
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer

# ─── Config ─────────────────────────────────────────────
UPLOAD_DIR = "./uploaded_excels"
os.makedirs(UPLOAD_DIR, exist_ok=True)

VECTOR_DIM = 384
INDEX_NAME = "core_agent_tool_index"
REDIS_URL = "redis://localhost:6379"
KEY_PREFIX = "core_agent:data:tool:"

# ─── RedisVL Setup ──────────────────────────────────────
redis_client = Redis.from_url(REDIS_URL, decode_responses=False)

model = SentenceTransformer("all-MiniLM-L6-v2")

def get_embedding(text: str) -> bytes:
    vector = model.encode(text)
    return np.array(vector, dtype=np.float32).tobytes()

# ─── Redis Schema ───────────────────────────────────────
schema = {
    "index": {"name": INDEX_NAME, "prefix": KEY_PREFIX},
    "fields": [
        {"name": "id", "type": "tag"},
        {"name": "name", "type": "text"},
        {"name": "type", "type": "tag"},
        {"name": "status", "type": "tag"},
        {"name": "location", "type": "text"},
        {"name": "quantity", "type": "numeric"},
        {"name": "unit", "type": "tag"},
        {"name": "weight", "type": "numeric"},
        {"name": "dim_length", "type": "numeric"},
        {"name": "dim_width", "type": "numeric"},
        {"name": "dim_height", "type": "numeric"},
        {"name": "created_at", "type": "text"},
        {"name": "updated_at", "type": "text"},
        {"name": "tags", "type": "tag"},
        {"name": "metadata", "type": "text"},
        {"name": "images", "type": "text"},
        {
            "name": "embedding",
            "type": "vector",
            "attrs": {
                "dims": VECTOR_DIM,
                "distance_metric": "cosine",
                "algorithm": "hnsw",
                "datatype": "float32"
            }
        }
    ]
}

index = SearchIndex.from_dict(schema, redis_client=redis_client)
if not index.exists():
    index.create(overwrite=False)

# ─── Router ─────────────────────────────────────────────
def build_upload_router(prefix: str = "/api") -> APIRouter:
    router = APIRouter(prefix=prefix)

    @router.post("/upload_tools_excel", summary="📄 Upload tools Excel và ghi đè RedisVL")
    async def upload_excel(file: UploadFile = File(...)):
        try:
            # Save file temporarily
            filename = f"{uuid.uuid4().hex}_{file.filename}"
            filepath = os.path.join(UPLOAD_DIR, filename)
            with open(filepath, "wb") as f:
                shutil.copyfileobj(file.file, f)

            # Load Excel
            wb = load_workbook(filename=filepath)
            sheet = wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            rows_processed = 0
            documents = []
            keys = []

            # Xóa toàn bộ doc key cùng prefix
            old_keys = redis_client.keys(f"{KEY_PREFIX}*")
            if old_keys:
                redis_client.delete(*old_keys)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                now = datetime.datetime.utcnow().isoformat()
                obj_id = data.get("id") or f"OBJ-{uuid.uuid4().hex[:6].upper()}"

                name = str(data.get("name", ""))
                type_ = str(data.get("type", ""))
                status = str(data.get("status", ""))
                metadata_raw = data.get("metadata") or "{}"
                metadata = json.loads(metadata_raw) if isinstance(metadata_raw, str) else metadata_raw
                metadata_str = json.dumps(metadata, ensure_ascii=False)

                # Embedding from combined fields
                embedding_text = f"{name} {type_} {status} {metadata_str}"
                embedding = get_embedding(embedding_text)

                doc = {
                    "id": obj_id,
                    "name": name,
                    "type": type_,
                    "status": status,
                    "location": str(data.get("location", "")),
                    "quantity": float(data.get("quantity") or 0),
                    "unit": str(data.get("unit", "")),
                    "weight": float(data.get("weight") or 0),
                    "dim_length": float(data.get("dimensions.length") or 0),
                    "dim_width": float(data.get("dimensions.width") or 0),
                    "dim_height": float(data.get("dimensions.height") or 0),
                    "created_at": data.get("created_at") or now,
                    "updated_at": data.get("updated_at") or now,
                    "tags": ",".join(str(data.get("tags", "")).split(",")),
                    "metadata": metadata_str,
                    "images": json.dumps(json.loads(data.get("images") or "[]"), ensure_ascii=False),
                    "embedding": embedding
                }

                redis_key = f"{KEY_PREFIX}{obj_id}"
                documents.append(doc)
                keys.append(redis_key)
                rows_processed += 1

            index.load(documents, keys=keys)

            return {
                "success": True,
                "message": f"✅ Uploaded {rows_processed} tools vào RedisVL",
                "file": filename,
                "key_prefix": KEY_PREFIX
            }

        except Exception as e:
            raise HTTPException(status_code=500, detail=f"❌ Failed to process Excel: {e}")

    return router
